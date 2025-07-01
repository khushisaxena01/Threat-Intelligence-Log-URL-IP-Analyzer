from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
import pandas as pd
import numpy as np
import os
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from multiprocessing import Pool, cpu_count
import json
import openpyxl
from openpyxl.styles import PatternFill
from werkzeug.utils import secure_filename
import io
import base64
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from collections import Counter
import re
from urllib.parse import urlparse
import ipaddress

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024 * 1024  # 5GB max file size
app.secret_key = 'your-secret-key-here'

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Global variables to store analysis results
analysis_results = {}
processing_status = {'status': 'idle', 'progress': 0, 'message': ''}

ALLOWED_EXTENSIONS = {'txt', 'csv', 'xlsx', 'xls', 'json'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

class LogAnalyzer:
    def __init__(self):
        self.logs_data = None
        self.urls_ips_data = None
        self.blocked_items = set()
        self.accepted_items = set()
        self.rejected_items = set()
        self.results = []
        self.hits = 0
        self.misses = 0
        self.processing_time = 0
        
    def load_file(self, filepath):
        """Load file based on extension with multiple format support"""
        try:
            ext = filepath.split('.')[-1].lower()
            
            if ext == 'csv':
                # Try different encodings and separators
                for encoding in ['utf-8', 'latin-1', 'cp1252']:
                    for sep in [',', ';', '\t', '|']:
                        try:
                            df = pd.read_csv(filepath, encoding=encoding, sep=sep)
                            if len(df.columns) > 1:
                                return df
                        except:
                            continue
                # Fallback
                return pd.read_csv(filepath, encoding='utf-8', sep=',')
                
            elif ext in ['xlsx', 'xls']:
                return pd.read_excel(filepath)
                
            elif ext == 'txt':
                # Try to read as structured data first
                try:
                    df = pd.read_csv(filepath, sep='\t', encoding='utf-8')
                    if len(df.columns) > 1:
                        return df
                except:
                    pass
                
                # Read as plain text and convert to DataFrame
                with open(filepath, 'r', encoding='utf-8') as f:
                    lines = [line.strip() for line in f.readlines() if line.strip()]
                return pd.DataFrame({'data': lines})
                
            elif ext == 'json':
                return pd.read_json(filepath)
                
        except Exception as e:
            print(f"Error loading file {filepath}: {str(e)}")
            return None
    
    def extract_urls_ips(self, text):
        """Extract URLs and IPs from text using regex"""
        urls = []
        ips = []
        
        # URL pattern
        url_pattern = r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
        urls.extend(re.findall(url_pattern, str(text)))
        
        # Domain pattern
        domain_pattern = r'(?:[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?\.)+[a-zA-Z]{2,}'
        domains = re.findall(domain_pattern, str(text))
        urls.extend([f"http://{domain}" for domain in domains])
        
        # IP pattern
        ip_pattern = r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b'
        ips.extend(re.findall(ip_pattern, str(text)))
        
        return urls, ips
    
    def categorize_urls_ips(self, df):
        """Categorize URLs/IPs into accepted, rejected, blocked"""
        global processing_status
        
        blocked = set()
        accepted = set()
        rejected = set()
        
        total_rows = len(df)
        
        for idx, row in df.iterrows():
            if idx % 100 == 0:  # Update progress every 100 rows
                processing_status['progress'] = int((idx / total_rows) * 30)  # 30% for categorization
                processing_status['message'] = f'Categorizing URLs/IPs: {idx}/{total_rows}'
            
            for col in row:
                if pd.isna(col):
                    continue
                    
                text = str(col).lower()
                
                # Check for status indicators
                if any(keyword in text for keyword in ['blocked', 'malicious', 'threat', 'dangerous', 'suspicious']):
                    urls, ips = self.extract_urls_ips(text)
                    blocked.update(urls + ips)
                elif any(keyword in text for keyword in ['accepted', 'safe', 'clean', 'allowed', 'whitelist']):
                    urls, ips = self.extract_urls_ips(text)
                    accepted.update(urls + ips)
                elif any(keyword in text for keyword in ['rejected', 'invalid', 'unreachable', 'failed']):
                    urls, ips = self.extract_urls_ips(text)
                    rejected.update(urls + ips)
                else:
                    # Try to extract URLs/IPs and categorize based on patterns
                    urls, ips = self.extract_urls_ips(text)
                    for item in urls + ips:
                        # Default categorization logic
                        if any(suspicious in item.lower() for suspicious in ['malware', 'phishing', 'spam', 'virus', 'trojan']):
                            blocked.add(item)
                        else:
                            accepted.add(item)
        
        return blocked, accepted, rejected
    
    def process_chunk(self, chunk_data):
        """Process a chunk of log data"""
        chunk_results = []
        chunk_hits = 0
        chunk_misses = 0
        
        logs_chunk, blocked_items = chunk_data
        
        for idx, row in logs_chunk.iterrows():
            hit_found = False
            matched_items = []
            
            for col in row:
                if pd.isna(col):
                    continue
                    
                text = str(col)
                urls, ips = self.extract_urls_ips(text)
                
                for item in urls + ips:
                    if item in blocked_items:
                        hit_found = True
                        matched_items.append(item)
            
            result = {
                'row_index': idx,
                'data': row.to_dict(),
                'status': 'HIT' if hit_found else 'MISS',
                'matched_items': matched_items
            }
            
            chunk_results.append(result)
            
            if hit_found:
                chunk_hits += 1
            else:
                chunk_misses += 1
        
        return chunk_results, chunk_hits, chunk_misses
    
    def analyze_logs(self, logs_file, urls_ips_file):
        """Main analysis function with multithreading"""
        global processing_status
        start_time = time.time()
        
        try:
            processing_status['status'] = 'processing'
            processing_status['progress'] = 0
            processing_status['message'] = 'Loading files...'
            
            # Load files
            self.logs_data = self.load_file(logs_file)
            self.urls_ips_data = self.load_file(urls_ips_file)
            
            if self.logs_data is None or self.urls_ips_data is None:
                raise Exception("Failed to load one or both files")
            
            processing_status['progress'] = 10
            processing_status['message'] = 'Categorizing URLs/IPs...'
            
            # Categorize URLs/IPs
            self.blocked_items, self.accepted_items, self.rejected_items = self.categorize_urls_ips(self.urls_ips_data)
            
            processing_status['progress'] = 40
            processing_status['message'] = 'Analyzing logs...'
            
            # Split logs into chunks for parallel processing
            num_threads = min(cpu_count(), 8)  # Limit to 8 threads max
            chunk_size = max(1, len(self.logs_data) // num_threads)
            chunks = []
            
            for i in range(0, len(self.logs_data), chunk_size):
                chunk = self.logs_data.iloc[i:i+chunk_size]
                chunks.append((chunk, self.blocked_items))
            
            # Process chunks in parallel
            self.results = []
            self.hits = 0
            self.misses = 0
            
            with ThreadPoolExecutor(max_workers=num_threads) as executor:
                future_to_chunk = {executor.submit(self.process_chunk, chunk): chunk for chunk in chunks}
                
                completed = 0
                for future in as_completed(future_to_chunk):
                    try:
                        chunk_results, chunk_hits, chunk_misses = future.result()
                        self.results.extend(chunk_results)
                        self.hits += chunk_hits
                        self.misses += chunk_misses
                        
                        completed += 1
                        progress = 40 + int((completed / len(chunks)) * 50)
                        processing_status['progress'] = progress
                        processing_status['message'] = f'Processing chunks: {completed}/{len(chunks)}'
                        
                    except Exception as e:
                        print(f"Error processing chunk: {str(e)}")
            
            processing_status['progress'] = 90
            processing_status['message'] = 'Generating reports...'
            
            self.processing_time = time.time() - start_time
            
            # Generate Excel report
            self.generate_excel_report()
            
            # Generate chart
            chart_data = self.generate_chart()
            
            processing_status['progress'] = 100
            processing_status['status'] = 'completed'
            processing_status['message'] = 'Analysis completed successfully!'
            
            return {
                'success': True,
                'hits': self.hits,
                'misses': self.misses,
                'total': len(self.results),
                'processing_time': round(self.processing_time, 2),
                'blocked_items_count': len(self.blocked_items),
                'chart_data': chart_data,
                'excel_file': 'analysis_results.xlsx'
            }
            
        except Exception as e:
            processing_status['status'] = 'error'
            processing_status['message'] = f'Error: {str(e)}'
            return {'success': False, 'error': str(e)}
    
    def generate_excel_report(self):
        """Generate Excel report with colored cells"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analysis Results"
        
        # Headers
        headers = ['Row Index', 'Status', 'Matched Items'] + list(self.logs_data.columns)
        ws.append(headers)
        
        # Color fills
        red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        
        # Add data
        for result in self.results:
            row_data = [
                result['row_index'],
                result['status'],
                ', '.join(result['matched_items']) if result['matched_items'] else 'None'
            ]
            
            # Add original row data
            for col in self.logs_data.columns:
                row_data.append(result['data'].get(col, ''))
            
            ws.append(row_data)
            
            # Apply coloring
            current_row = ws.max_row
            fill = red_fill if result['status'] == 'HIT' else green_fill
            
            for cell in ws[current_row]:
                cell.fill = fill
        
        # Save file
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'analysis_results.xlsx')
        wb.save(filepath)
    
    def generate_chart(self):
        """Generate pie chart data"""
        total = self.hits + self.misses
        if total == 0:
            return None
        
        hit_percentage = (self.hits / total) * 100
        miss_percentage = (self.misses / total) * 100
        
        return {
            'labels': ['Risky (HIT)', 'Safe (MISS)'],
            'data': [self.hits, self.misses],
            'percentages': [round(hit_percentage, 2), round(miss_percentage, 2)],
            'colors': ['#FF6B6B', '#4ECDC4']
        }

# Initialize analyzer
analyzer = LogAnalyzer()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    global processing_status
    
    try:
        # Reset processing status
        processing_status = {'status': 'idle', 'progress': 0, 'message': ''}
        
        # Check if files are present
        if 'logs_file' not in request.files or 'urls_ips_file' not in request.files:
            return jsonify({'success': False, 'error': 'Both files are required'})
        
        logs_file = request.files['logs_file']
        urls_ips_file = request.files['urls_ips_file']
        
        if logs_file.filename == '' or urls_ips_file.filename == '':
            return jsonify({'success': False, 'error': 'Please select both files'})
        
        if not (allowed_file(logs_file.filename) and allowed_file(urls_ips_file.filename)):
            return jsonify({'success': False, 'error': 'Invalid file format'})
        
        # Save files
        logs_filename = secure_filename(logs_file.filename)
        urls_ips_filename = secure_filename(urls_ips_file.filename)
        
        logs_path = os.path.join(app.config['UPLOAD_FOLDER'], logs_filename)
        urls_ips_path = os.path.join(app.config['UPLOAD_FOLDER'], urls_ips_filename)
        
        logs_file.save(logs_path)
        urls_ips_file.save(urls_ips_path)
        
        # Start analysis in background thread
        def run_analysis():
            global analysis_results
            analysis_results = analyzer.analyze_logs(logs_path, urls_ips_path)
        
        analysis_thread = threading.Thread(target=run_analysis)
        analysis_thread.start()
        
        return jsonify({'success': True, 'message': 'Files uploaded successfully. Analysis started.'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/progress')
def get_progress():
    return jsonify(processing_status)

@app.route('/results')
def results():
    global analysis_results
    if analysis_results and analysis_results.get('success'):
        return render_template('results.html', results=analysis_results)
    else:
        return redirect(url_for('index'))

@app.route('/download_excel')
def download_excel():
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'analysis_results.xlsx')
        return send_file(filepath, as_attachment=True, download_name='analysis_results.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)})

if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5000)
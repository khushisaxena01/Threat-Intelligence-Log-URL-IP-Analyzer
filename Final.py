import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np
import threading
import time
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import seaborn as sns
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
import ipaddress
from urllib.parse import urlparse
import tldextract
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import json
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

class ThreatAnalyzerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Threat Intelligence Log Analyzer")
        self.root.geometry("1200x800")
        self.root.configure(bg='#f0f0f0')
        
        # Initialize variables
        self.log_file_path = None
        self.threat_file_path = None
        self.results = []
        self.threat_urls = set()
        self.threat_ips = set()
        self.blocked_items = set()
        self.processing_time = 0
        self.total_items = 0
        self.hits = 0
        self.misses = 0
        
        # Create GUI elements
        self.create_widgets()
        
    def create_widgets(self):
        """Create and arrange GUI widgets"""
        
        # Main title
        title_label = tk.Label(self.root, text="Threat Intelligence Log Analyzer", 
                              font=("Arial", 20, "bold"), bg='#f0f0f0', fg='#2c3e50')
        title_label.pack(pady=20)
        
        # File selection frame
        file_frame = tk.Frame(self.root, bg='#f0f0f0')
        file_frame.pack(pady=10, padx=20, fill='x')
        
        # Log file selection
        log_frame = tk.Frame(file_frame, bg='#f0f0f0')
        log_frame.pack(fill='x', pady=5)
        
        tk.Label(log_frame, text="Select Log File:", font=("Arial", 12, "bold"), 
                bg='#f0f0f0').pack(side='left', padx=(0, 10))
        
        self.log_file_label = tk.Label(log_frame, text="No file selected", 
                                      font=("Arial", 10), bg='#f0f0f0', fg='#7f8c8d')
        self.log_file_label.pack(side='left', padx=(0, 10))
        
        log_browse_btn = tk.Button(log_frame, text="Browse Log File", 
                                  command=self.browse_log_file,
                                  bg='#3498db', fg='white', font=("Arial", 10, "bold"),
                                  padx=20, pady=5)
        log_browse_btn.pack(side='right')
        
        # Threat intelligence file selection
        threat_frame = tk.Frame(file_frame, bg='#f0f0f0')
        threat_frame.pack(fill='x', pady=5)
        
        tk.Label(threat_frame, text="Select Threat Intelligence File:", 
                font=("Arial", 12, "bold"), bg='#f0f0f0').pack(side='left', padx=(0, 10))
        
        self.threat_file_label = tk.Label(threat_frame, text="No file selected", 
                                         font=("Arial", 10), bg='#f0f0f0', fg='#7f8c8d')
        self.threat_file_label.pack(side='left', padx=(0, 10))
        
        threat_browse_btn = tk.Button(threat_frame, text="Browse Threat File", 
                                     command=self.browse_threat_file,
                                     bg='#e74c3c', fg='white', font=("Arial", 10, "bold"),
                                     padx=20, pady=5)
        threat_browse_btn.pack(side='right')
        
        # Analyze button
        self.analyze_btn = tk.Button(self.root, text="🔍 ANALYZE LOGS", 
                                    command=self.start_analysis,
                                    bg='#27ae60', fg='white', font=("Arial", 14, "bold"),
                                    padx=40, pady=10, state='disabled')
        self.analyze_btn.pack(pady=20)
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.root, variable=self.progress_var, 
                                           maximum=100, length=400)
        self.progress_bar.pack(pady=10)
        
        # Status label
        self.status_label = tk.Label(self.root, text="Please select both log and threat files", 
                                    font=("Arial", 10), bg='#f0f0f0', fg='#7f8c8d')
        self.status_label.pack(pady=5)
        
        # Results frame
        self.results_frame = tk.Frame(self.root, bg='#f0f0f0')
        self.results_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.results_frame)
        self.notebook.pack(fill='both', expand=True)
        
        # Summary tab
        self.summary_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.summary_frame, text="📊 Summary")
        
        # Visualization tab
        self.viz_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.viz_frame, text="📈 Visualization")
        
        # Details tab
        self.details_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.details_frame, text="📋 Details")
        
    def browse_log_file(self):
        """Browse and select log file"""
        file_path = filedialog.askopenfilename(
            title="Select Log File",
            filetypes=[
                ("All Supported", "*.csv;*.xlsx;*.txt;*.json"),
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx"),
                ("Text files", "*.txt"),
                ("JSON files", "*.json"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            self.log_file_path = file_path
            self.log_file_label.config(text=f"Selected: {os.path.basename(file_path)}", fg='#27ae60')
            self.check_files_selected()
    
    def browse_threat_file(self):
        """Browse and select threat intelligence file"""
        file_path = filedialog.askopenfilename(
            title="Select Threat Intelligence File",
            filetypes=[
                ("All Supported", "*.csv;*.xlsx;*.txt;*.json"),
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx"),
                ("Text files", "*.txt"),
                ("JSON files", "*.json"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            self.threat_file_path = file_path
            self.threat_file_label.config(text=f"Selected: {os.path.basename(file_path)}", fg='#27ae60')
            self.check_files_selected()
    
    def check_files_selected(self):
        """Check if both files are selected and enable analyze button"""
        if self.log_file_path and self.threat_file_path:
            self.analyze_btn.config(state='normal')
            self.status_label.config(text="Ready to analyze! Click 'ANALYZE LOGS' button.", fg='#27ae60')
        else:
            self.analyze_btn.config(state='disabled')
    
    def load_file_data(self, file_path):
        """Load data from various file formats"""
        try:
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.csv':
                df = pd.read_csv(file_path, encoding='utf-8', low_memory=False)
            elif file_ext == '.xlsx':
                df = pd.read_excel(file_path)
            elif file_ext == '.txt':
                with open(file_path, 'r', encoding='utf-8') as f:
                    lines = [line.strip() for line in f if line.strip()]
                df = pd.DataFrame(lines, columns=['data'])
            elif file_ext == '.json':
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if isinstance(data, list):
                    df = pd.DataFrame(data)
                else:
                    df = pd.DataFrame([data])
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            return df
            
        except Exception as e:
            raise Exception(f"Error loading file {file_path}: {str(e)}")
    
    def extract_urls_ips(self, df):
        """Extract URLs and IPs from dataframe"""
        urls_ips = []
        
        for _, row in df.iterrows():
            for cell in row:
                if pd.notna(cell):
                    cell_str = str(cell).strip()
                    if cell_str:
                        urls_ips.append(cell_str)
        
        return urls_ips
    
    def classify_threat_item(self, item):
        """Classify threat intelligence item as ACCEPTED, REJECTED, or BLOCKED"""
        try:
            # Check if it's a URL
            if self.is_url(item):
                # For threat intelligence, we'll mark suspicious/malicious URLs as BLOCKED
                if self.is_suspicious_url(item):
                    return 'BLOCKED'
                else:
                    return 'ACCEPTED'
            
            # Check if it's an IP
            elif self.is_ip(item):
                # For threat intelligence, we'll mark private/reserved IPs as ACCEPTED
                # and suspicious IPs as BLOCKED
                try:
                    ip_obj = ipaddress.ip_address(item)
                    if ip_obj.is_private or ip_obj.is_loopback:
                        return 'ACCEPTED'
                    elif self.is_suspicious_ip(item):
                        return 'BLOCKED'
                    else:
                        return 'ACCEPTED'
                except:
                    return 'REJECTED'
            
            else:
                return 'REJECTED'
                
        except Exception:
            return 'REJECTED'
    
    def is_url(self, text):
        """Check if text is a valid URL"""
        try:
            # Add protocol if missing
            if not text.startswith(('http://', 'https://', 'ftp://')):
                text = 'http://' + text
            
            result = urlparse(text)
            return all([result.scheme, result.netloc])
        except:
            return False
    
    def is_ip(self, text):
        """Check if text is a valid IP address"""
        try:
            ipaddress.ip_address(text)
            return True
        except:
            return False
    
    def is_suspicious_url(self, url):
        """Check if URL has suspicious characteristics"""
        suspicious_patterns = [
            r'malware', r'phishing', r'spam', r'virus', r'trojan',
            r'suspicious', r'malicious', r'blocked', r'dangerous',
            r'\d{1,3}-\d{1,3}-\d{1,3}-\d{1,3}',  # IP-like domains
            r'[a-z0-9]{20,}',  # Very long random strings
        ]
        
        for pattern in suspicious_patterns:
            if re.search(pattern, url, re.IGNORECASE):
                return True
        return False
    
    def is_suspicious_ip(self, ip):
        """Check if IP has suspicious characteristics"""
        try:
            ip_obj = ipaddress.ip_address(ip)
            # Mark certain ranges as suspicious (this is just for demo)
            suspicious_ranges = [
                ipaddress.ip_network('185.0.0.0/8'),
                ipaddress.ip_network('188.0.0.0/8'),
            ]
            
            for network in suspicious_ranges:
                if ip_obj in network:
                    return True
            return False
        except:
            return False
    
    def process_chunk(self, log_chunk):
        """Process a chunk of log data"""
        chunk_results = []
        
        for log_item in log_chunk:
            log_item = str(log_item).strip()
            if not log_item:
                continue
            
            # Check against blocked items
            is_hit = False
            matched_threat = None
            
            for blocked_item in self.blocked_items:
                if self.items_match(log_item, blocked_item):
                    is_hit = True
                    matched_threat = blocked_item
                    break
            
            result = {
                'log_item': log_item,
                'status': 'HIT' if is_hit else 'MISS',
                'matched_threat': matched_threat if is_hit else '',
                'item_type': 'URL' if self.is_url(log_item) else 'IP' if self.is_ip(log_item) else 'OTHER',
                'timestamp': datetime.now().isoformat()
            }
            
            chunk_results.append(result)
        
        return chunk_results
    
    def items_match(self, log_item, threat_item):
        """Check if log item matches threat item"""
        # Simple string matching (can be enhanced with fuzzy matching)
        log_item_clean = log_item.lower().strip()
        threat_item_clean = threat_item.lower().strip()
        
        # Exact match
        if log_item_clean == threat_item_clean:
            return True
        
        # Domain matching for URLs
        if self.is_url(log_item) and self.is_url(threat_item):
            try:
                log_domain = tldextract.extract(log_item).registered_domain
                threat_domain = tldextract.extract(threat_item).registered_domain
                return log_domain == threat_domain
            except:
                pass
        
        # Substring matching
        if threat_item_clean in log_item_clean or log_item_clean in threat_item_clean:
            return True
        
        return False
    
    def start_analysis(self):
        """Start the analysis in a separate thread"""
        self.analyze_btn.config(state='disabled')
        self.progress_var.set(0)
        self.status_label.config(text="Starting analysis...", fg='#f39c12')
        
        # Start analysis in separate thread
        analysis_thread = threading.Thread(target=self.run_analysis)
        analysis_thread.daemon = True
        analysis_thread.start()
    
    def run_analysis(self):
        """Run the complete analysis"""
        try:
            start_time = time.time()
            
            # Load threat intelligence
            self.update_status("Loading threat intelligence file...")
            self.progress_var.set(10)
            threat_df = self.load_file_data(self.threat_file_path)
            threat_items = self.extract_urls_ips(threat_df)
            
            # Classify threat items and extract blocked ones
            self.update_status("Classifying threat intelligence...")
            self.progress_var.set(20)
            
            for item in threat_items:
                classification = self.classify_threat_item(item)
                if classification == 'BLOCKED':
                    self.blocked_items.add(item.lower().strip())
            
            # Load log file
            self.update_status("Loading log file...")
            self.progress_var.set(30)
            log_df = self.load_file_data(self.log_file_path)
            log_items = self.extract_urls_ips(log_df)
            
            self.total_items = len(log_items)
            
            # Process logs using multithreading
            self.update_status("Analyzing logs with multithreading...")
            self.progress_var.set(40)
            
            # Split log items into chunks for parallel processing
            max_workers = min(32, (self.total_items // 1000) + 1)  # Dynamic worker count
            chunk_size = max(1, self.total_items // max_workers)
            chunks = [log_items[i:i + chunk_size] for i in range(0, len(log_items), chunk_size)]
            
            self.results = []
            completed_chunks = 0
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_chunk = {executor.submit(self.process_chunk, chunk): chunk for chunk in chunks}
                
                for future in as_completed(future_to_chunk):
                    chunk_results = future.result()
                    self.results.extend(chunk_results)
                    completed_chunks += 1
                    progress = 40 + (completed_chunks / len(chunks)) * 40
                    self.progress_var.set(progress)
            
            # Calculate statistics
            self.update_status("Calculating statistics...")
            self.progress_var.set(85)
            
            self.hits = sum(1 for r in self.results if r['status'] == 'HIT')
            self.misses = self.total_items - self.hits
            self.processing_time = time.time() - start_time
            
            # Generate Excel report
            self.update_status("Generating Excel report...")
            self.progress_var.set(90)
            self.generate_excel_report()
            
            # Update GUI with results
            self.update_status("Updating display...")
            self.progress_var.set(95)
            self.root.after(0, self.display_results)
            
            self.progress_var.set(100)
            self.update_status(f"Analysis completed in {self.processing_time:.2f} seconds!")
            
        except Exception as e:
            self.update_status(f"Error during analysis: {str(e)}")
            messagebox.showerror("Analysis Error", f"An error occurred during analysis:\n{str(e)}")
        finally:
            self.root.after(0, lambda: self.analyze_btn.config(state='normal'))
    
    def update_status(self, message):
        """Update status label thread-safely"""
        self.root.after(0, lambda: self.status_label.config(text=message))
    
    def generate_excel_report(self):
        """Generate Excel report with color coding"""
        try:
            # Create DataFrame
            df = pd.DataFrame(self.results)
            
            # Create Excel writer
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"threat_analysis_report_{timestamp}.xlsx"
            
            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Analysis Results', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Analysis Results']
                
                # Define fills
                red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                
                # Apply color formatting
                for row in range(2, len(df) + 2):  # Skip header row
                    status_cell = worksheet[f'B{row}']  # Status column
                    if status_cell.value == 'HIT':
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = red_fill
                    else:
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = green_fill
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # Add summary sheet
                summary_data = [
                    ['Metric', 'Value'],
                    ['Total Items Analyzed', self.total_items],
                    ['Hits (Risky)', self.hits],
                    ['Misses (Safe)', self.misses],
                    ['Hit Percentage', f"{(self.hits/self.total_items)*100:.2f}%" if self.total_items > 0 else "0%"],
                    ['Processing Time (seconds)', f"{self.processing_time:.2f}"],
                    ['Blocked Threats Found', len(self.blocked_items)],
                    ['Analysis Date', datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
                ]
                
                summary_df = pd.DataFrame(summary_data[1:], columns=summary_data[0])
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            self.excel_filename = excel_filename
            
        except Exception as e:
            print(f"Error generating Excel report: {e}")
    
    def display_results(self):
        """Display analysis results in GUI"""
        # Clear previous results
        for widget in self.summary_frame.winfo_children():
            widget.destroy()
        for widget in self.viz_frame.winfo_children():
            widget.destroy()
        for widget in self.details_frame.winfo_children():
            widget.destroy()
        
        # Summary tab
        summary_text = f"""
        📊 ANALYSIS SUMMARY
        
        Total Items Analyzed: {self.total_items:,}
        
        🔴 HITS (Risky): {self.hits:,} ({(self.hits/self.total_items)*100:.2f}%)
        🟢 MISSES (Safe): {self.misses:,} ({(self.misses/self.total_items)*100:.2f}%)
        
        ⏱️ Processing Time: {self.processing_time:.2f} seconds
        🛡️ Blocked Threats in Database: {len(self.blocked_items):,}
        
        📁 Excel Report Generated: {getattr(self, 'excel_filename', 'Not generated')}
        """
        
        summary_label = tk.Label(self.summary_frame, text=summary_text, 
                                font=("Courier", 12), justify='left', bg='white')
        summary_label.pack(padx=20, pady=20, anchor='w')
        
        # Show blocked items if any hits
        if self.hits > 0:
            blocked_items_text = "\n🚨 DETECTED THREATS:\n"
            hit_results = [r for r in self.results if r['status'] == 'HIT']
            for i, result in enumerate(hit_results[:10]):  # Show first 10
                blocked_items_text += f"{i+1}. {result['log_item']} → {result['matched_threat']}\n"
            
            if len(hit_results) > 10:
                blocked_items_text += f"... and {len(hit_results) - 10} more threats detected.\n"
            
            blocked_label = tk.Label(self.summary_frame, text=blocked_items_text, 
                                   font=("Courier", 10), justify='left', bg='white', fg='red')
            blocked_label.pack(padx=20, pady=10, anchor='w')
        
        # Visualization tab
        self.create_visualization()
        
        # Details tab
        self.create_details_view()
    
    def create_visualization(self):
        """Create pie chart visualization"""
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 6))
        fig.patch.set_facecolor('white')
        
        # Pie chart
        labels = ['Safe (MISS)', 'Risky (HIT)']
        sizes = [self.misses, self.hits]
        colors = ['#2ecc71', '#e74c3c']
        explode = (0.05, 0.05)
        
        ax1.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90, 
                colors=colors, explode=explode, shadow=True)
        ax1.set_title('Security Analysis Results', fontsize=14, fontweight='bold')
        
        # Bar chart
        categories = ['Safe\n(MISS)', 'Risky\n(HIT)']
        values = [self.misses, self.hits]
        bars = ax2.bar(categories, values, color=colors, alpha=0.8)
        
        # Add value labels on bars
        for bar, value in zip(bars, values):
            height = bar.get_height()
            ax2.text(bar.get_x() + bar.get_width()/2., height + max(values)*0.01,
                    f'{value:,}', ha='center', va='bottom', fontweight='bold')
        
        ax2.set_title('Item Count Comparison', fontsize=14, fontweight='bold')
        ax2.set_ylabel('Count')
        
        plt.tight_layout()
        
        # Embed in tkinter
        canvas = FigureCanvasTkAgg(fig, self.viz_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True, padx=10, pady=10)
    
    def create_details_view(self):
        """Create detailed results view"""
        # Create frame with scrollbar
        details_container = tk.Frame(self.details_frame)
        details_container.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Create treeview for results
        columns = ('Item', 'Status', 'Type', 'Matched Threat')
        tree = ttk.Treeview(details_container, columns=columns, show='headings', height=15)
        
        # Define headings
        tree.heading('Item', text='Log Item')
        tree.heading('Status', text='Status')
        tree.heading('Type', text='Type')
        tree.heading('Matched Threat', text='Matched Threat')
        
        # Configure column widths
        tree.column('Item', width=300)
        tree.column('Status', width=100)
        tree.column('Type', width=100)
        tree.column('Matched Threat', width=200)
        
        # Add scrollbars
        v_scrollbar = ttk.Scrollbar(details_container, orient='vertical', command=tree.yview)
        h_scrollbar = ttk.Scrollbar(details_container, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack treeview and scrollbars
        tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        details_container.grid_rowconfigure(0, weight=1)
        details_container.grid_columnconfigure(0, weight=1)
        
        # Populate treeview
        for result in self.results[:1000]:  # Limit to first 1000 for performance
            values = (
                result['log_item'][:50] + '...' if len(result['log_item']) > 50 else result['log_item'],
                result['status'],
                result['item_type'],
                result['matched_threat'][:50] + '...' if len(result['matched_threat']) > 50 else result['matched_threat']
            )
            
            item_id = tree.insert('', 'end', values=values)
            
            # Color code the rows
            if result['status'] == 'HIT':
                tree.set(item_id, 'Status', '🔴 HIT')
            else:
                tree.set(item_id, 'Status', '🟢 MISS')

def main():
    """Main function to run the GUI application"""
    root = tk.Tk()
    app = ThreatAnalyzerGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()